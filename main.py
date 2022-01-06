import time
import webbrowser
import datetime
import openpyxl
import keyboard


# gets current weekday as an integer
def current_weekday():
    datetime.datetime.today()
    return datetime.datetime.today().weekday()


# kill window
def kill_window():
    keyboard.press_and_release('ctrl+w')


# main code
def main_function(weekday, workbook):
    sheet_list = workbook.sheetnames
    sheet = workbook[sheet_list[weekday]]
    start = False
    for i in range(1, (sheet.max_row + 1)):
        while True:
            if not start:
                if (datetime.datetime.now().hour == sheet.cell(row=i, column=2)) and (
                        datetime.datetime.now().minute == sheet.cell(row=i, column=3)):
                    webbrowser.open(sheet.cell(row=i, column=5))
                    start = True
            else:
                time.sleep((sheet.cell(row=i, column=4)) * 60)
                kill_window()
                break


if __name__ == "__main__":
    # load workbook and sheet names
    path = 'C:/Users/dsaum/Downloads/CourseData.xlsx'
    wb = openpyxl.load_workbook(path)
    wd = current_weekday()
    main_function(wd, wb)
